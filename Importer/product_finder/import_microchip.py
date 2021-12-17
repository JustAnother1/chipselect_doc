#!/usr/bin/python3
# -*- coding: utf-8 -*-

import sys
import openpyxl

import ucdb_sql

# check file Headers
headerRow = str(1)
fileHeaders = {'Product' : 'A', 
'Status' : 'B',
'Packages' : 'AX',
'CPU Type' : 'G',
'Max CPU Speed (MHz)' : 'H',
'Program Memory Size (KB)' : 'N',
'SRAM (Bytes)' : 'P',
'Operation Voltage Min (V)' : 'J',
'Operation Voltage Max (V)' : 'K',
'Temp Range Max' : 'L',
}

def verifyDocument(sheet):
    if None == sheet:
        return False
    # scan the header row and if the cell contains one of the headers (fileHeaders.key()) then that row is the value for that header
    for col in range(1, 255):
        c = sheet.cell(row=int(headerRow), column=col)
        text = c.value
        if text in fileHeaders.keys():
            fileHeaders[text] = c.column_letter

    for header in fileHeaders.keys():
        if header != sheet[fileHeaders[header] + headerRow].value:
            print("File does not look like it is supposed to be!")
            print(fileHeaders[header] + headerRow + " : " + sheet[fileHeaders[header] + headerRow].value + " expected : " + header)
            return False
    return True

def getDeviceFromRow(sheet, row):
    row = str(row)
    dev = {}
    # name
    name = sheet[fileHeaders['Product'] + row].value
    if None == name :
        return None
    if 1 > len(name) :
        return None
    dev['name'] = name
    # MarketState
    state = sheet[fileHeaders['Status'] + row].value
    if "-" == state :
        pass
    else:
        dev['MarketState'] = state
    # Package
    pack = sheet[fileHeaders['Packages'] + row].value
    if "-" == pack :
        pass
    else:
        dev['Package'] = pack
    # Architecture
    arch = sheet[fileHeaders['CPU Type'] + row].value
    if "-" == arch :
        pass
    else:
        dev['Architecture'] = arch
    # CPU max clock
    cpuClock = sheet[fileHeaders['Max CPU Speed (MHz)'] + row].value
    if "-" == cpuClock :
        pass
    else:
        dev['CPU_clock_max_MHz'] = cpuClock
    # Flash size
    flash = sheet[fileHeaders['Program Memory Size (KB)'] + row].value
    if "-" == flash :
        pass
    else:
        dev['Flash_size_kB'] = flash
    # RAM size
    ram = sheet[fileHeaders['SRAM (Bytes)'] + row].value
    if "-" == ram :
        pass
    else:
        dev['RAM_size_kB'] = ram/1024
    # Vmin
    vmin = sheet[fileHeaders['Operation Voltage Min (V)'] + row].value
    if "-" == vmin :
        pass
    else:
        dev['Supply_Voltage_min_V'] = vmin
    # Vmax
    vmax = sheet[fileHeaders['Operation Voltage Max (V)'] + row].value
    if "-" == vmax :
        pass
    else:
        dev['Supply_Voltage_max_V'] = vmax

    maxTemp = sheet[fileHeaders['Temp Range Max'] + row].value
    # sometimes it is just "-"
    if "-" == maxTemp :
        pass
    else:
        maxTemp = str(maxTemp)
        nums = maxTemp.split(',')
        max = 0
        for n in nums:
            if "-" == n :
                pass
            if int(n) > int(max):
                max = int(n)
        dev['Operating_Temperature_max_degC'] = max
    return dev



def importExcelFile(fname):
    print('now importing the file ' + fname)
    wb = openpyxl.load_workbook(fname)
    print("sheets: " + str(wb.sheetnames))
    sheet = wb.active
    if False == verifyDocument(sheet):
        sys.exit(1)

    curRow = 2
    while True:
        print("Importing line " + str(curRow))
        device = getDeviceFromRow(sheet, curRow)
        if None == device:
            print("No more devices in the file!")
            break

        device['Vendor'] = 'Microchip Technology Inc.'
        
        if False == ucdb_sql.addDeviceToDatabase(device):
            print("Failed to store device into the data base!")
            break
        curRow = curRow + 1

#ucdb_sql.connect2localDB()
ucdb_sql.connect2remoteDB()
importExcelFile('microcontroller_and_processors-2020-08-14.xlsx')
ucdb_sql.closeDB()
print("Done!")
