#!/usr/bin/python3
# -*- coding: utf-8 -*-

import sys
import openpyxl

import ucdb_sql

# check file Headers
headerRow = str(9)
fileHeaders = {'Parts' : 'A', 
'Status' : 'D',
'Package Type' : 'AX',  
'Core Type' : 'F',
'Operating Frequency [Max] (MHz)' : 'G',
'Flash (kB)' : 'H', 
'SRAM (kB)' : 'I', 
'Supply Voltage [Min to Max] (V)' : 'X', 
'Ambient Operating Temperature (Min to Max) (℃)' : 'Y', 
}

def verifyDocument(sheet):
    if None == sheet:
        return False
    # scan the header row and if the cell contains one of the headers (fileHeaders.key()) then that row is the value for that header
    for col in range(1, 255):
        c = sheet.cell(row=int(headerRow), column=col)
        text = c.value
        if text in fileHeaders.keys():
            fileHeaders[text] = c.column_letter

    for header in fileHeaders.keys():
        if header != sheet[fileHeaders[header] + headerRow].value:
            print("File does not look like it is supposed to be!")
            print(fileHeaders[header] + headerRow + " : " + sheet[fileHeaders[header] + headerRow].value + " expected : " + header)
            return False
    return True

def getDeviceFromRow(sheet, row):
    row = str(row)
    dev = {}
    # name
    name = sheet[fileHeaders['Parts'] + row].value
    if None == name :
        return None
    if 1 > len(name) :
        return None
    dev['name'] = name
    # MarketState
    state = sheet[fileHeaders['Status'] + row].value
    if "-" == state or None == state :
        pass
    else:
        dev['MarketState'] = state
    # Package
    pack = sheet[fileHeaders['Package Type'] + row].value
    if "-" == pack or None == pack :
        pass
    else:
        dev['Package'] = pack
    # Architecture
    arch = sheet[fileHeaders['Core Type'] + row].value
    if "-" == arch or None == arch :
        pass
    else:
        dev['Architecture'] = arch
    # CPU max clock
    cpuClock = sheet[fileHeaders['Operating Frequency [Max] (MHz)'] + row].value
    if "-" == cpuClock or None == cpuClock :
        pass
    else:
        dev['CPU_clock_max_MHz'] = cpuClock
    # Flash size
    flash = sheet[fileHeaders['Flash (kB)'] + row].value
    if "-" == flash or None == flash:
        pass
    else:
        dev['Flash_size_kB'] = flash
    # RAM size
    ram = sheet[fileHeaders['SRAM (kB)'] + row].value
    if "-" == ram or None == ram :
        pass
    else:
        dev['RAM_size_kB'] = ram

    v_range = sheet[fileHeaders['Supply Voltage [Min to Max] (V)'] + row].value
    if None != v_range :
        if 2 < len(v_range):
            v_range = v_range.split('to')
            
            # Vmin
            vmin = v_range[0].strip()
            if "-" == vmin :
                pass
            else:
                dev['Supply_Voltage_min_V'] = vmin
            # Vmax
            vmax = v_range[1].strip()
            if "-" == vmax :
                pass
            else:
                dev['Supply_Voltage_max_V'] = vmax

    temp_range = sheet[fileHeaders['Ambient Operating Temperature (Min to Max) (℃)'] + row].value
    if None != temp_range :
        if 2 < len(temp_range):
            temp_range = temp_range.split('to')

            minTemp = temp_range[0].strip()
            dev['Operating_Temperature_min_degC'] = minTemp

            maxTemp = temp_range[1].strip()
            dev['Operating_Temperature_max_degC'] = maxTemp
    return dev



def importExcelFile(fname):
    print('now importing the file ' + fname)
    wb = openpyxl.load_workbook(fname)
    print("sheets: " + str(wb.sheetnames))
    sheet = wb.active
    if False == verifyDocument(sheet):
        sys.exit(1)

    curRow = 10 # first row that contains a device
    while True:
        print("Importing line " + str(curRow))
        device = getDeviceFromRow(sheet, curRow)
        if None == device:
            print("No more devices in the file!")
            break

        device['Vendor'] = 'NXP Semiconductors'
        
        if False == ucdb_sql.addDeviceToDatabase(device):
            print("Failed to store device into the data base!")
            break
        curRow = curRow + 1

#ucdb_sql.connect2localDB()
ucdb_sql.connect2remoteDB()
importExcelFile('NXPProductSelectorResults.xlsx')
ucdb_sql.closeDB()
print("Done!")
