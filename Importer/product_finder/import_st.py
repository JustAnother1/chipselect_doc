#!/usr/bin/python3
# -*- coding: utf-8 -*-

import sys
import openpyxl

import ucdb_sql

# check file Headers
headerRow = str(6)
fileHeaders = {'Part Number' : 'A', 
'Marketing Status' : 'C',
'Package' : 'D',
'Core' : 'E',
'Operating Frequency (MHz)' : 'F',
'Flash Size (kB) (Prog)' : 'H',
'RAM Size (kB)' : 'J',
'Supply Voltage (V) min' : 'AH',
'Supply Voltage (V) max' : 'AI',
'Operating Temperature (°C) min' : 'AL',
'Operating Temperature (°C) max' : 'AM',
}

def verifyDocument(sheet):
    if None == sheet:
        return False
    # scan the header row and if the cell contains one of the headers (fileHeaders.key()) then that row is the value for that header
    for col in range(1, 255):
        c = sheet.cell(row=int(headerRow), column=col)
        text = c.value
        if text in fileHeaders.keys():
            fileHeaders[text] = c.column_letter

    for header in fileHeaders.keys():
        if header != sheet[fileHeaders[header] + headerRow].value:
            print("File does not look like it is supposed to be!")
            print(fileHeaders[header] + headerRow + " : " + sheet[fileHeaders[header] + headerRow].value + " expected : " + header)
            return False
    return True

def getDeviceFromRow(sheet, row):
    row = str(row)
    dev = {}
    # name
    name = sheet[fileHeaders['Part Number'] + row].value
    if None == name :
        return None
    if 1 > len(name) :
        return None
    dev['name'] = name
    # MarketState
    state = sheet[fileHeaders['Marketing Status'] + row].value
    if "-" == state :
        pass
    else:
        dev['MarketState'] = state
    # Package
    pack = sheet[fileHeaders['Package'] + row].value
    if "-" == pack :
        pass
    else:
        dev['Package'] = pack
    # Architecture
    arch = sheet[fileHeaders['Core'] + row].value
    if "-" == arch :
        pass
    else:
        dev['Architecture'] = arch
    # CPU max clock
    cpuClock = sheet[fileHeaders['Operating Frequency (MHz)'] + row].value
    if "-" == cpuClock :
        pass
    else:
        dev['CPU_clock_max_MHz'] = cpuClock
    # Flash size
    flash = sheet[fileHeaders['Flash Size (kB) (Prog)'] + row].value
    if "-" == flash :
        pass
    else:
        dev['Flash_size_kB'] = flash
    # RAM size
    ram = sheet[fileHeaders['RAM Size (kB)'] + row].value
    if "-" == ram :
        pass
    else:
        dev['RAM_size_kB'] = ram
    # Vmin
    vmin = sheet[fileHeaders['Supply Voltage (V) min'] + row].value
    if "-" == vmin :
        pass
    else:
        dev['Supply_Voltage_min_V'] = vmin
    # Vmax
    vmax = sheet[fileHeaders['Supply Voltage (V) max'] + row].value
    if "-" == vmax :
        pass
    else:
        dev['Supply_Voltage_max_V'] = vmax
    # sometimes ST hast two values in this 
    minTemp = sheet[fileHeaders['Operating Temperature (°C) min'] + row].value
    # sometimes it is just "-"
    if "-" == minTemp :
        pass
    else:
        nums = minTemp.split(',')
        min = 9000
        for n in nums:
            if "-" == n :
                pass
            if int(n) < int(min):
                min = int(n)
        dev['Operating_Temperature_min_degC'] = min
    # sometimes ST hast two values in this "105,85"
    maxTemp = sheet[fileHeaders['Operating Temperature (°C) max'] + row].value
    # sometimes it is just "-"
    if "-" == maxTemp :
        pass
    else:
        nums = maxTemp.split(',')
        max = 0
        for n in nums:
            if "-" == n :
                pass
            if int(n) > int(max):
                max = int(n)
        dev['Operating_Temperature_max_degC'] = max
    return dev



def importExcelFile(fname):
    print('now importing the file ' + fname)
    wb = openpyxl.load_workbook(fname)
    print("sheets: " + str(wb.sheetnames))
    #sheet = wb.active
    sheet = wb['ProductsList']
    if False == verifyDocument(sheet):
        sys.exit(1)

    curRow = 8
    while True:
        print("Importing line " + str(curRow))
        device = getDeviceFromRow(sheet, curRow)
        if None == device:
            print("No more devices in the file!")
            break

        device['Vendor'] = 'STMicroelectronics'
        
        if False == ucdb_sql.addDeviceToDatabase(device):
            print("Failed to store device into the data base!")
            break
        curRow = curRow + 1

#ucdb_sql.connect2localDB()
ucdb_sql.connect2remoteDB()
importExcelFile('productslist_mainstream.xlsx')
importExcelFile('productslist_low_power.xlsx') 
importExcelFile('productslist_high_performance.xlsx') 
ucdb_sql.closeDB()
print("Done!")
